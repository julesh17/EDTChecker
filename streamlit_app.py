import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime, date, time
from dateutil import parser as dtparser
from openpyxl import load_workbook

# ---------------- Utilities ----------------

def normalize_group_label(x):
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    s = str(x).strip()
    if not s:
        return None
    m = re.search(r'G\s*\.?\s*(\d+)', s, re.I)
    if m:
        return f'G {m.group(1)}'
    m2 = re.search(r'^(?:groupe)?\s*(\d+)$', s, re.I)
    if m2:
        return f'G {m2.group(1)}'
    return s

def is_time_like(x):
    if x is None:
        return False
    if isinstance(x, (pd.Timestamp, datetime, time)):
        return True
    s = str(x).strip()
    if not s:
        return False
    if re.match(r'^\d{1,2}[:hH]\d{2}(\s*[AaPp][Mm]\.?)?$', s):
        return True
    return False

def to_time(x):
    if x is None:
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().time()
    if isinstance(x, datetime):
        return x.time()
    s = str(x).strip()
    if not s:
        return None
    s2 = s.replace('h', ':').replace('H', ':')
    try:
        dt = dtparser.parse(s2, dayfirst=True)
        return dt.time()
    except Exception:
        return None

def to_date(x):
    if x is None:
        return None
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().date()
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    if not s:
        return None
    try:
        dt = dtparser.parse(s, dayfirst=True, fuzzy=True)
        return dt.date()
    except Exception:
        return None

def find_week_rows(df):
    rows = []
    for i in range(len(df)):
        try:
            v = df.iat[i, 0]
        except Exception:
            v = None
        if isinstance(v, str) and re.match(r'^\s*S\s*\d+', v.strip(), re.I):
            rows.append(i)
    return rows

def find_slot_rows(df):
    rows = []
    for i in range(len(df)):
        try:
            v = df.iat[i, 0]
        except Exception:
            v = None
        if isinstance(v, str) and re.match(r'^\s*H\s*\d+', v.strip(), re.I):
            rows.append(i)
    return rows

# ---------------- Helpers Fusion ----------------

def get_merged_map_from_bytes(xls_bytes, sheet_name):
    """
    Retourne un dict {(row_zero_based, col_zero_based): (r1,c1,r2,c2)} pour les ranges fusionnées
    basé sur le sheet_name. On lit depuis bytes (io.BytesIO).
    """
    wb = load_workbook(io.BytesIO(xls_bytes), data_only=True)
    ws = wb[sheet_name]
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        r1, r2 = merged.min_row, merged.max_row
        c1, c2 = merged.min_col, merged.max_col
        # convert to zero-based indices
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                merged_map[(r - 1, c - 1)] = (r1 - 1, c1 - 1, r2 - 1, c2 - 1)
    return merged_map

# ---------------- Parsing ----------------

def parse_sheet_to_events(uploaded_file, sheet_name):
    """
    uploaded_file : object retourné par st.file_uploader (file-like)
    sheet_name : nom de la feuille
    """
    # read bytes once and reuse for both pandas and openpyxl
    if hasattr(uploaded_file, 'read'):
        xls_bytes = uploaded_file.read()
        # reset pointer for uploaded_file if needed by streamlit later (UploadedFile cannot be rewound easily),
        # but we've already consumed it; downstream code uses xls (pd.ExcelFile) created below from bytes.
    else:
        # uploaded_file might be a path string
        with open(uploaded_file, 'rb') as f:
            xls_bytes = f.read()

    # read sheet into dataframe
    try:
        df = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=sheet_name, header=None, engine='openpyxl')
    except Exception as e:
        # in case of error, return empty
        return []

    nrows, ncols = df.shape

    # build merged map for this sheet
    try:
        merged_map = get_merged_map_from_bytes(xls_bytes, sheet_name)
    except Exception:
        merged_map = {}

    s_rows = find_week_rows(df)
    h_rows = find_slot_rows(df)

    raw_events = []

    for r in h_rows:
        p_candidates = [s for s in s_rows if s <= r]
        if not p_candidates:
            continue
        p = max(p_candidates)
        date_row = p + 1
        group_row = p + 2

        date_cols = [c for c in range(ncols) if date_row < nrows and to_date(df.iat[date_row, c]) is not None]

        for c in date_cols:
            for col in (c, c + 1):
                if col >= ncols:
                    continue
                try:
                    summary = df.iat[r, col]
                except Exception:
                    summary = None
                if pd.isna(summary) or summary is None:
                    continue
                summary_str = str(summary).strip()
                if not summary_str:
                    continue

                # collect teachers (multiple)
                teachers = []
                if (r + 2) < nrows:
                    for off in range(2, 6):
                        if (r + off) >= nrows:
                            break
                        try:
                            t = df.iat[r + off, col]
                        except Exception:
                            t = None
                        if t and not pd.isna(t) and not is_time_like(t):
                            s_t = str(t).strip()
                            if s_t:
                                teachers.append(s_t)
                teachers = list(dict.fromkeys(teachers))

                # find first time-like cell after summary
                stop_idx = None
                for off in range(1, 12):
                    idx = r + off
                    if idx >= nrows:
                        break
                    try:
                        if is_time_like(df.iat[idx, col]):
                            stop_idx = idx
                            break
                    except Exception:
                        continue
                if stop_idx is None:
                    stop_idx = min(r + 7, nrows)

                # description cells
                desc_parts = []
                for idx in range(r + 1, stop_idx):
                    if idx >= nrows:
                        break
                    try:
                        cell = df.iat[idx, col]
                    except Exception:
                        cell = None
                    if pd.isna(cell) or cell is None:
                        continue
                    s = str(cell).strip()
                    if not s:
                        continue
                    if to_date(cell) is not None:
                        continue
                    if s in teachers or s == summary_str:
                        continue
                    desc_parts.append(s)
                desc_text = " | ".join(dict.fromkeys(desc_parts))

                # start/end time
                start_val = None
                end_val = None
                for off in range(1, 13):
                    idx = r + off
                    if idx >= nrows:
                        break
                    try:
                        v = df.iat[idx, col]
                    except Exception:
                        v = None
                    if is_time_like(v):
                        if start_val is None:
                            start_val = v
                        elif end_val is None and v != start_val:
                            end_val = v
                            break
                if start_val is None or end_val is None:
                    continue
                start_t = to_time(start_val)
                end_t = to_time(end_val)
                if start_t is None or end_t is None:
                    continue

                # day date
                date_cell = df.iat[date_row, c]
                d = to_date(date_cell)
                if d is None:
                    continue

                dtstart = datetime.combine(d, start_t)
                dtend = datetime.combine(d, end_t)

                # groups
                gl = None
                gl_next = None
                if group_row < nrows:
                    try:
                        gl_raw = df.iat[group_row, col]
                        gl = normalize_group_label(gl_raw)
                    except Exception:
                        gl = None
                if (col + 1) < ncols and group_row < nrows:
                    try:
                        gl_next_raw = df.iat[group_row, col + 1]
                        gl_next = normalize_group_label(gl_next_raw)
                    except Exception:
                        gl_next = None

                is_left_col = (col == c)
                right_summary = None
                if (col + 1) < ncols:
                    try:
                        right_summary = df.iat[r, col + 1]
                    except Exception:
                        right_summary = None

                groups = set()
                # ---- Use merged_map to detect real Excel merge between (r,col) and (r,col+1) ----
                if is_left_col:
                    merged_here = merged_map.get((r, col))
                    merged_right = merged_map.get((r, col + 1))
                    # if both cells belong to a same merged range that spans the adjacent column,
                    # treat it as a shared course (G1+G2)
                    if merged_here is not None and merged_right is not None and merged_here == merged_right:
                        # add both group labels if present
                        if gl:
                            groups.add(gl)
                        if gl_next:
                            groups.add(gl_next)
                    else:
                        # not merged across columns: normal behavior (only left group)
                        if gl:
                            groups.add(gl)
                else:
                    # right column event (only add its group)
                    if gl:
                        groups.add(gl)

                raw_events.append({
                    'summary': summary_str,
                    'teachers': set(teachers),
                    'descriptions': set([desc_text]) if desc_text else set(),
                    'start': dtstart,
                    'end': dtend,
                    'groups': groups
                })

    # merge raw events by (summary, start, end)
    merged = {}
    for e in raw_events:
        key = (e['summary'], e['start'], e['end'])
        if key not in merged:
            merged[key] = {
                'summary': e['summary'],
                'teachers': set(),
                'descriptions': set(),
                'start': e['start'],
                'end': e['end'],
                'groups': set()
            }
        merged[key]['teachers'].update(e.get('teachers', set()))
        merged[key]['descriptions'].update(e.get('descriptions', set()))
        merged[key]['groups'].update(e.get('groups', set()))

    out = []
    for v in merged.values():
        out.append({
            'summary': v['summary'],
            'teachers': sorted([t for t in v['teachers'] if t and str(t).lower() not in ['nan','none']]),
            'description': " | ".join(sorted([d for d in v['descriptions'] if d and d.strip()])),
            'start': v['start'],
            'end': v['end'],
            'groups': sorted(list(v['groups']))
        })
    return out

# ---------------- Maquette loader ----------------

def read_maquette(xls):
    """
    Lit la feuille 'Maquette' en essayant de trouver les colonnes sujets (C) et total (M).
    Retourne un DataFrame ordonné (même ordre que dans la feuille) avec colonnes:
      ['subject', 'target', 'promo'] (promo peut être empty string).
    Gestion robuste si les noms de colonnes varient : on tente de détecter par noms,
    sinon on tombe sur lecture par index de colonnes (C -> idx 2, M -> idx 12).
    """
    # détecte la feuille maquette
    sheet_candidates = [s for s in xls.sheet_names if s.lower() == 'maquette' or 'maquette' in s.lower()]
    if not sheet_candidates:
        return pd.DataFrame(columns=['subject','target','promo'])

    sheet = sheet_candidates[0]

    # tente lecture standard (avec header)
    try:
        mq = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
    except Exception:
        # fallback lecture sans header
        try:
            mq = pd.read_excel(xls, sheet_name=sheet, header=None, engine='openpyxl')
        except Exception:
            return pd.DataFrame(columns=['subject','target','promo'])

    # si on a des noms de colonnes explicites, on essaie de les utiliser
    if mq.shape[0] == 0:
        return pd.DataFrame(columns=['subject','target','promo'])

    # Normalize column lookup (case-insensitive)
    cols = {str(c).strip().lower(): c for c in mq.columns}

    # Cherche colonne sujet (préférence sur mots-clés)
    subject_col = None
    for k in cols:
        if any(w in k for w in ['mati', 'subject', 'module', 'ue', 'course']):
            subject_col = cols[k]; break

    # cherche colonne target (total heures)
    target_col = None
    for k in cols:
        if any(w in k for w in ['total', 'seanc', 'session', 'target', 'cibl', 'nb', 'heure', 'hours']):
            target_col = cols[k]; break

    # cherche promo (si présente)
    promo_col = None
    for k in cols:
        if any(w in k for w in ['promo','promotion','year','p1','p2']):
            promo_col = cols[k]; break

    # Si on a pas trouvé subject / target via noms, on tente lecture par index (C -> idx 2, M -> idx 12)
    if subject_col is None or target_col is None:
        # relire sans header pour utiliser indices si possible
        try:
            mq2 = pd.read_excel(xls, sheet_name=sheet, header=None, engine='openpyxl')
        except Exception:
            mq2 = None

        if mq2 is not None and mq2.shape[1] > 12:
            # colonne C (index 2) et M (index 12)
            subject_col_idx = 2
            target_col_idx = 12
            rows = []
            for i in range(len(mq2)):
                subj = mq2.iat[i, subject_col_idx]
                targ = mq2.iat[i, target_col_idx]
                if pd.isna(subj) or str(subj).strip() == '':
                    continue
                try:
                    tval = float(targ) if not pd.isna(targ) else None
                except Exception:
                    tval = None
                rows.append({'subject': str(subj).strip(), 'target': tval, 'promo': ''})
            return pd.DataFrame(rows, columns=['subject','target','promo'])
        else:
            # si impossible, retourne DataFrame vide structurée
            return pd.DataFrame(columns=['subject','target','promo'])

    # si on arrive ici, on a trouvé subject_col et target_col par noms
    rows = []
    for _, row in mq.iterrows():
        subject = row.get(subject_col)
        if pd.isna(subject) or str(subject).strip() == '':
            continue
        promo = row.get(promo_col) if promo_col else ''
        target = row.get(target_col) if target_col else None
        try:
            tval = float(target) if target is not None and not pd.isna(target) else None
        except Exception:
            tval = None
        promo_str = str(promo).strip() if promo and not pd.isna(promo) else ''
        rows.append({'subject': str(subject).strip(), 'target': tval, 'promo': promo_str})
    return pd.DataFrame(rows, columns=['subject','target','promo'])

# ---------------- Aggregation helpers ----------------

def build_events_index(xls, sheet_names):
    out = {}
    for s in sheet_names:
        try:
            events = parse_sheet_to_events(xls, s)
            out[s] = events
        except Exception:
            out[s] = []
    return out

def sum_hours_by_subject(events):
    totals = {}
    for ev in events:
        if ev['start'] and ev['end']:
            delta = (ev['end'] - ev['start']).total_seconds() / 3600.0
        else:
            delta = 0
        subj = ev['summary']
        totals[subj] = totals.get(subj, 0) + delta
    return totals

def flatten_events_table(events):
    rows = []
    for ev in events:
        rows.append({
            'subject': ev['summary'],
            'start': ev['start'],
            'end': ev['end'],
            'groups': ', '.join(ev['groups']) if ev['groups'] else '',
            'teachers': ', '.join(ev['teachers']) if ev['teachers'] else '',
            'description': ev['description'] if ev.get('description') else ''
        })
    if rows:
        return pd.DataFrame(rows)
    else:
        return pd.DataFrame(columns=['subject','start','end','groups','teachers','description'])

# ---------------- UI ----------------

st.set_page_config(page_title='EDT Reports', layout='wide')
st.title('EDT — Rapports et comparaisons')

uploaded = st.file_uploader('Charger le fichier Excel (EDT + Maquette)', type=['xlsx'])
if uploaded is None:
    st.info('Upload un fichier .xlsx pour commencer.')
    st.stop()

try:
    # create a pandas ExcelFile from uploaded bytes so read_maquette / listing work
    xls_bytes_for_listing = uploaded.read()
    xls = pd.ExcelFile(io.BytesIO(xls_bytes_for_listing), engine='openpyxl')
    sheets = xls.sheet_names
except Exception as e:
    st.error('Impossible de lire le fichier Excel: ' + str(e))
    st.stop()

st.write('Feuilles trouvées :', sheets)

promo_sheets = [s for s in sheets if s.strip().upper() in ['EDT P1','EDT P2','P1','P2','EDT P1 ','EDT P2 ']]
if not promo_sheets:
    promo_sheets = [s for s in sheets if 'EDT' in s.upper() or 'P1' in s.upper() or 'P2' in s.upper()]

promo_sheets = [s for s in ['EDT P1','EDT P2'] if s in sheets] or promo_sheets

# build events using the original uploaded file object — parse_sheet_to_events expects the uploaded file-like,
# but it already consumed uploaded.read() above, so we pass the bytes we saved to parse_sheet_to_events:
class _BytesWrapper:
    def __init__(self, b):
        self._b = b
    def read(self):
        return self._b

uploaded_bytes_wrapper = _BytesWrapper(xls_bytes_for_listing)

events_by_promo = build_events_index(uploaded_bytes_wrapper, promo_sheets)
maquette_df = read_maquette(xls)

# navigation
page = st.sidebar.selectbox('Page', [
    '1 — Comparaison avec Maquette',
    '2 — Récap par matière',
    '3 — Récap par enseignant',
    '4 — Récap complet textuel',
    '5 — Récap heures par type d’enseignant'
])

# ---------- Page 1 ----------
if page.startswith('1'):
    st.header('Comparaison des heures par matière vs Maquette')

    if maquette_df.empty:
        st.warning('Feuille Maquette non trouvée ou non lisible — impossible de faire un comparatif complet.')
        for promo, evs in events_by_promo.items():
            st.subheader(promo)
            totals = sum_hours_by_subject(evs)
            dfp = pd.DataFrame([(k, v) for k, v in totals.items()], columns=['subject','hours'])
            if dfp.empty:
                st.info('Aucune séance détectée sur cette feuille.')
            else:
                st.dataframe(dfp.sort_values('subject').reset_index(drop=True))
        st.stop()

    # choix promo / groupe
    promo_choice = st.selectbox('Sélectionner la promo (feuille)', options=promo_sheets)
    group_choice = st.selectbox('Sélectionner le groupe', options=['G 1', 'G 2'], index=0)

    # normalisation du choix de groupe
    def parse_group_sel(sel):
        s = sel.strip().upper().replace(' ', '')
        if s in ['G1','G 1']:
            return {'G 1','G1'}
        if s in ['G2','G 2']:
            return {'G 2','G2'}
        return {sel}

    sel_groups = parse_group_sel(group_choice)

    # événements pour la promo choisie
    evs = events_by_promo.get(promo_choice, [])

    # matières à ignorer
    IGNORE_SUBJECTS = {
        "erasmus day",
        "forum international",
        "période entreprise",
        "férié",
        "mission à l'international",
        "matière",
        "matières",
        "divers"
    }

    # calcul heures + séances par matière (filtrées par groupe et liste noire)
    def sum_hours_by_subject_and_group(events, groups_filter):
        totals = {}
        counts = {}
        for ev in events:
            subj = ev['summary']
            if subj and subj.strip().lower() in IGNORE_SUBJECTS:
                continue  # skip

            # filtrage par groupe
            ev_groups_norm = set([g.strip().upper().replace(' ', '') for g in ev.get('groups', [])])
            if groups_filter is None:
                matches = True
            else:
                matches = len(ev_groups_norm.intersection(
                    {g.strip().upper().replace(' ', '') for g in groups_filter}
                )) > 0
            if not matches:
                continue

            # durée
            if ev['start'] and ev['end']:
                delta = (ev['end'] - ev['start']).total_seconds() / 3600.0
            else:
                delta = 0
            totals[subj] = totals.get(subj, 0) + delta
            counts[subj] = counts.get(subj, 0) + 1
        return totals, counts

    totals_by_subject, counts_by_subject = sum_hours_by_subject_and_group(evs, sel_groups)

    # tableau final basé sur la maquette (dans l’ordre de la maquette)
    rows = []
    for _, row in maquette_df.iterrows():
        subj = str(row['subject']).strip()
        if subj.lower() in IGNORE_SUBJECTS:
            continue  # skip

        target = row['target'] if 'target' in row else None
        hours = totals_by_subject.get(subj, 0.0)
        sessions = counts_by_subject.get(subj, 0)
        diff = None
        if target is not None and not pd.isna(target):
            diff = hours - float(target)
        rows.append({
            'subject': subj,
            'target_hours': target,
            'entered_hours': round(hours, 2),
            'diff_hours': round(diff, 2) if diff is not None else None,
            'sessions_entered': sessions
        })

    out_df = pd.DataFrame(rows, columns=['subject','target_hours','entered_hours','diff_hours','sessions_entered'])

    st.markdown(f"### Résultats pour la feuille **{promo_choice}** — groupe **{group_choice}**")
    st.write("Toutes les matières de la maquette sont affichées (ordre maquette). Les matières ignorées ne figurent pas dans ce tableau.")

    def highlight_row(r):
        if r['target_hours'] is None or pd.isna(r['target_hours']):
            return ['background-color: #fff3cd']*len(r)
        if r['diff_hours'] is not None and abs(r['diff_hours']) > 0.001:
            return ['background-color: #f8d7da']*len(r)
        return ['']*len(r)

    try:
        st.dataframe(out_df.style.apply(highlight_row, axis=1))
    except Exception:
        st.dataframe(out_df)

    # résumé
    total_expected = out_df['target_hours'].dropna().sum() if 'target_hours' in out_df.columns else None
    total_entered = out_df['entered_hours'].sum()
    st.markdown("**Résumé**")
    if total_expected is not None:
        st.write(f"- Heures attendues (maquette): **{total_expected:.2f} h**")
    else:
        st.write("- Heures attendues : **non disponibles**")
    st.write(f"- Heures saisies: **{total_entered:.2f} h**")
    if total_expected is not None:
        st.write(f"- Écart total: **{(total_entered - total_expected):+.2f} h**")

    st.markdown("**Vérifications rapides**")
    st.write("- Jaune = pas d'heures attendues dans la maquette.")
    st.write("- Rouge = écart entre heures saisies et maquette.")

# ---------- Page 2 ----------
elif page.startswith('2'):
    st.header('Récapitulatif par matière (sélectionne une matière)')
    subjects = set()
    for evs in events_by_promo.values():
        for ev in evs:
            subjects.add(ev['summary'])
    subjects = sorted(list(subjects))
    if not subjects:
        st.warning('Aucune matière trouvée.')
    else:
        chosen = st.selectbox('Choisir une matière', options=subjects)
        st.write(f'Séances pour : **{chosen}**')
        for promo, evs in events_by_promo.items():
            df_ev = flatten_events_table([e for e in evs if e['summary'] == chosen])
            st.subheader(promo)
            if df_ev.empty:
                st.info('Aucune séance pour cette matière.')
            else:
                st.dataframe(df_ev.sort_values('start').reset_index(drop=True))

# ---------- Page 3 ----------
elif page.startswith('3'):
    st.header('Récapitulatif par enseignant')
    teachers = set()
    for evs in events_by_promo.values():
        for ev in evs:
            for t in ev['teachers']:
                if t and t.lower() not in ['nan','none']:
                    teachers.add(t)
    teachers = sorted(list(teachers))
    if not teachers:
        st.warning('Aucun enseignant détecté.')
    else:
        chosen = st.selectbox('Choisir un enseignant', options=teachers)
        st.write(f'Séances pour : **{chosen}**')
        for promo, evs in events_by_promo.items():
            df_ev = flatten_events_table([e for e in evs if chosen in e['teachers']])
            st.subheader(promo)
            if df_ev.empty:
                st.info('Aucune séance pour cet enseignant dans cette promo.')
            else:
                st.dataframe(df_ev.sort_values('start').reset_index(drop=True))

# ---------- Page 4 ----------
elif page.startswith('4'):
    st.header('Récapitulatif textuel complet (par promo)')
    for promo, evs in events_by_promo.items():
        st.subheader(promo)
        if not evs:
            st.info('Aucune séance détectée.')
            continue
        by_subject = {}
        for e in evs:
            by_subject.setdefault(e['summary'], []).append(e)
        for subj in sorted(by_subject.keys()):
            st.markdown(f'**{subj}**')
            lst = sorted(by_subject[subj], key=lambda x: x['start'])
            for e in lst:
                groups = ', '.join(e['groups']) if e['groups'] else '—'
                teachers = ', '.join(e['teachers']) if e['teachers'] else '—'
                desc = e['description'] if e.get('description') else ''
                st.write(
                    f"- {e['start'].strftime('%Y-%m-%d %H:%M')} → {e['end'].strftime('%H:%M')} "
                    f"— Groupes: {groups} — Enseignant(s): {teachers} {('- '+desc) if desc else ''}"
                )

# ---------- Page 5 ----------
elif page.startswith('5'):
    st.header("Récap heures par type d’enseignant")

    # Charger la feuille enseignants
    try:
        enseignants_df = pd.read_excel(xls, sheet_name="Enseignants", engine="openpyxl")
    except Exception as e:
        st.error("Impossible de lire la feuille 'Enseignants': " + str(e))
        st.stop()

    # Normaliser noms enseignants et types
    enseignants_map = {}
    for _, row in enseignants_df.iterrows():
        nom = str(row.iloc[0]).strip()  # colonne A = nom enseignant
        type_str = str(row.iloc[2]).strip().upper() if len(row) > 2 else ""
        if nom and nom.lower() not in ["nan","none"]:
            enseignants_map[nom] = type_str

    def classify_event(ev):
        """Retourne le type principal de l'événement"""
        if not ev['teachers']:
            return "Autonomie"
        types = []
        for t in ev['teachers']:
            t_norm = str(t).strip()
            type_str = enseignants_map.get(t_norm, "")
            if "CESI" in type_str:
                types.append("CESI")
            elif "UPS TOULOUSE III" in type_str:
                types.append("UPS TOULOUSE III")
            else:
                types.append("Non CESI")
        # Si plusieurs profs, on garde tous (concaténés)
        return set(types)

    def hours(ev):
        if ev['start'] and ev['end']:
            return (ev['end'] - ev['start']).total_seconds()/3600.0
        return 0.0

    # Table de résultats
    results = []
    for promo, evs in events_by_promo.items():
        counters = {"Autonomie":0.0, "CESI":0.0, "Non CESI":0.0, "UPS TOULOUSE III":0.0}
        for ev in evs:
            cat = classify_event(ev)
            if isinstance(cat, str):
                counters[cat] += hours(ev)
            else:  # plusieurs types (co-enseignement)
                h = hours(ev) / len(cat)
                for c in cat:
                    counters[c] += h
        results.append({"Promo": promo, **{k: round(v,2) for k,v in counters.items()}})

    # Ajout global (toutes promos)
    total_counters = {"Autonomie":0.0, "CESI":0.0, "Non CESI":0.0, "UPS TOULOUSE III":0.0}
    for r in results:
        for k in total_counters:
            total_counters[k] += r[k]
    results.append({"Promo": "Toutes promos", **{k: round(v,2) for k,v in total_counters.items()}})

    df_res = pd.DataFrame(results)
    st.dataframe(df_res)

